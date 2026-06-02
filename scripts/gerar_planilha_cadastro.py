"""
Gera planilha Excel modelo para cadastro do banco de dados Projetista V2.
"""
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1

# ── Cores ──────────────────────────────────────────────────────────────────
COR_HEADER_ESCURO  = "4A235A"   # roxo escuro
COR_HEADER_MEDIO   = "7B2D8B"   # roxo médio
COR_HEADER_CLARO   = "F3E5F5"   # lilás claro
COR_OBRIGATORIO    = "FFF9C4"   # amarelo claro  → campo obrigatório
COR_OPCIONAL       = "F1F8E9"   # verde claro    → campo opcional
COR_AUTO           = "E3F2FD"   # azul claro     → gerado automaticamente
COR_EXEMPLO        = "E8F5E9"   # verde suave    → linha de exemplo
COR_INSTRUCAO      = "FFF3E0"   # laranja suave  → linha de instrução
COR_REFERENCIA     = "EDE7F6"   # lavanda        → chave estrangeira
COR_BRANCO         = "FFFFFF"
COR_CINZA_LINHA    = "FAFAFA"

def borda_fina():
    lado = Side(style="thin", color="CCCCCC")
    return Border(left=lado, right=lado, top=lado, bottom=lado)

def borda_media():
    lado = Side(style="medium", color="7B2D8B")
    return Border(left=lado, right=lado, top=lado, bottom=lado)

def celula(ws, linha, col, valor, negrito=False, cor_fundo=None,
           cor_texto="000000", alinhamento="left", tamanho=10,
           borda=True, italico=False, wrap=False):
    c = ws.cell(row=linha, column=col, value=valor)
    c.font = Font(name="Arial", bold=negrito, italic=italico,
                  color=cor_texto, size=tamanho)
    if cor_fundo:
        c.fill = PatternFill("solid", fgColor=cor_fundo)
    c.alignment = Alignment(horizontal=alinhamento, vertical="center",
                             wrap_text=wrap)
    if borda:
        c.border = borda_fina()
    return c

def linha_header_principal(ws, linha, titulo, ncols, cor=COR_HEADER_ESCURO):
    ws.merge_cells(start_row=linha, start_column=1,
                   end_row=linha, end_column=ncols)
    c = ws.cell(row=linha, column=1, value=titulo)
    c.font = Font(name="Arial", bold=True, color="FFFFFF", size=13)
    c.fill = PatternFill("solid", fgColor=cor)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = borda_media()
    ws.row_dimensions[linha].height = 28

def linha_header_col(ws, linha, colunas):
    """colunas: lista de (nome, largura, cor_fundo)"""
    for i, (nome, largura, cor) in enumerate(colunas, 1):
        c = celula(ws, linha, i, nome, negrito=True,
                   cor_fundo=cor, cor_texto="FFFFFF" if cor == COR_HEADER_MEDIO else "333333",
                   alinhamento="center", tamanho=9)
        ws.column_dimensions[get_column_letter(i)].width = largura
    ws.row_dimensions[linha].height = 22

def linha_legenda(ws, linha, ncols):
    ws.merge_cells(start_row=linha, start_column=1,
                   end_row=linha, end_column=ncols)
    c = ws.cell(row=linha, column=1,
                value="🟡 Obrigatório   🟢 Opcional   🔵 Auto (não preencher)   🟣 Chave estrangeira (ver aba IDs)")
    c.font = Font(name="Arial", italic=True, size=9, color="555555")
    c.fill = PatternFill("solid", fgColor="FFFDE7")
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[linha].height = 18

def linha_dados(ws, linha, valores, cor_fundo=COR_BRANCO):
    cor = COR_CINZA_LINHA if linha % 2 == 0 else COR_BRANCO
    if cor_fundo != COR_BRANCO:
        cor = cor_fundo
    for i, v in enumerate(valores, 1):
        celula(ws, linha, i, v, cor_fundo=cor)

# ══════════════════════════════════════════════════════════════════════════════
wb = Workbook()
wb.remove(wb.active)

# ══════════════════════════════════════════════════════════════════════════════
# ABA 0 — INSTRUÇÕES
# ══════════════════════════════════════════════════════════════════════════════
ws_inst = wb.create_sheet("📋 INSTRUÇÕES")
ws_inst.sheet_properties.tabColor = "4A235A"
ws_inst.column_dimensions["A"].width = 80
ws_inst.row_dimensions[1].height = 50

ws_inst.merge_cells("A1:A1")
c = ws_inst["A1"]
c.value = "PROJETISTA V2 — Planilha de Cadastro do Banco de Dados"
c.font = Font(name="Arial", bold=True, size=16, color="FFFFFF")
c.fill = PatternFill("solid", fgColor="4A235A")
c.alignment = Alignment(horizontal="center", vertical="center")

instrucoes = [
    ("", None),
    ("📌 COMO USAR ESTA PLANILHA", "7B2D8B"),
    ("", None),
    ("1. Preencha cada aba na ORDEM indicada (1→10). Algumas tabelas dependem de outras.", None),
    ("2. Não altere os cabeçalhos das colunas.", None),
    ("3. Campos 🔵 AUTOMÁTICO não precisam ser preenchidos (ID gerado pelo banco).", None),
    ("4. Campos 🟡 OBRIGATÓRIO devem ser sempre preenchidos.", None),
    ("5. Campos 🟣 CHAVE ESTRANGEIRA — consulte a aba '🔑 IDs' para ver os IDs corretos.", None),
    ("6. Após preencher, use o script de importação para carregar os dados no banco.", None),
    ("", None),
    ("📋 ORDEM DAS ABAS", "7B2D8B"),
    ("", None),
    ("  1. 📏 Unidades de Medida    → Sem dependências. Preencha primeiro.", None),
    ("  2. 🏭 Fabricantes           → Sem dependências.", None),
    ("  3. 📂 Categorias            → Sem dependências.", None),
    ("  4. 🍖 Tipos de Produto      → Sem dependências.", None),
    ("  5. 🌡️ Perfis de Produto     → Depende de: Tipos de Produto", None),
    ("  6. ⚙️ Equipamentos          → Depende de: Categorias, Fabricantes, Unidades", None),
    ("  7. 📊 Performance Equip.    → Depende de: Equipamentos  ← DADOS DO FABRICANTE", None),
    ("  8. 🔧 Materiais             → Depende de: Categorias, Fabricantes, Unidades", None),
    ("  9. 🔩 Componentes           → Depende de: Categorias, Fabricantes", None),
    (" 10. 📈 Performance Comp.     → Depende de: Componentes", None),
    ("", None),
    ("💡 DICAS", "7B2D8B"),
    ("", None),
    ("  • Temperatura de evaporação: use -10, -15, -20, -25, -30 °C", None),
    ("  • Temperatura de condensação: padrão 45°C (ajuste conforme projeto)", None),
    ("  • Capacidade em kcal/h: retirar dos catálogos técnicos dos fabricantes", None),
    ("  • Fluidos aceitos: R404A, R290, R22, R448A, R407C, R134a", None),
    ("  • Delta T padrão para evaporadoras: 8°C", None),
    ("  • Para compressores: qtde_ventiladores=0, diametro=0, vazao=0, flecha=0", None),
]

for i, (texto, cor) in enumerate(instrucoes, 2):
    ws_inst.row_dimensions[i].height = 20
    c = ws_inst.cell(row=i, column=1, value=texto)
    if cor:
        c.font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=cor)
    else:
        c.font = Font(name="Arial", size=10, color="333333")
    c.alignment = Alignment(horizontal="left", vertical="center")

# ══════════════════════════════════════════════════════════════════════════════
# ABA IDs — tabela de referência de IDs
# ══════════════════════════════════════════════════════════════════════════════
ws_ids = wb.create_sheet("🔑 IDs (Referência)")
ws_ids.sheet_properties.tabColor = "1565C0"

linha_header_principal(ws_ids, 1, "IDs DE REFERÊNCIA — Consulte antes de preencher as outras abas", 6, "1565C0")

secoes = [
    ("UNIDADES DE MEDIDA", [
        ("1", "unidade", "un"),
        ("2", "metro", "m"),
        ("3", "metro quadrado", "m²"),
        ("4", "quilograma", "kg"),
        ("5", "litro", "L"),
        ("6", "conjunto", "cj"),
    ]),
    ("CATEGORIAS", [
        ("1", "Condensadora", "Equipamento"),
        ("2", "Evaporadora", "Equipamento"),
        ("3", "Compressor", "Equipamento"),
        ("4", "Válvula de Expansão", "Componente"),
        ("5", "Filtro Secador", "Componente"),
        ("6", "Visor de Líquido", "Componente"),
        ("7", "Válvula Solenoide", "Componente"),
        ("8", "Pressostato", "Componente"),
        ("9", "Tubulação de Cobre", "Material"),
        ("10", "Isolamento Térmico", "Material"),
        ("11", "Material Elétrico", "Material"),
        ("12", "Painel Frigorífico", "Material"),
        ("13", "Solda e Fluxo", "Material"),
    ]),
    ("FABRICANTES", [
        ("1", "Tecumseh", ""),
        ("2", "Embraco", ""),
        ("3", "Bitzer", ""),
        ("4", "Danfoss", ""),
        ("5", "Parker", ""),
        ("6", "Elgin", ""),
        ("7", "Genérico", ""),
    ]),
    ("TIPOS DE PRODUTO", [
        ("1", "Carnes e Aves", ""),
        ("2", "Laticínios", ""),
        ("3", "FLV (Frutas, Legumes e Verduras)", ""),
        ("4", "Pescados", ""),
        ("5", "Frios e Embutidos", ""),
        ("6", "Sorvetes e Congelados", ""),
        ("7", "Bebidas", ""),
        ("8", "Padaria e Confeitaria", ""),
        ("9", "Geral / Industrial", ""),
    ]),
]

linha_atual = 3
for titulo, dados in secoes:
    ws_ids.merge_cells(start_row=linha_atual, start_column=1,
                       end_row=linha_atual, end_column=3)
    c = ws_ids.cell(row=linha_atual, column=1, value=titulo)
    c.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=COR_HEADER_MEDIO)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws_ids.row_dimensions[linha_atual].height = 20
    linha_atual += 1
    for id_val, nome, extra in dados:
        ws_ids.cell(row=linha_atual, column=1, value=int(id_val)).font = Font(name="Arial", bold=True, size=10)
        ws_ids.cell(row=linha_atual, column=2, value=nome).font = Font(name="Arial", size=10)
        ws_ids.cell(row=linha_atual, column=3, value=extra).font = Font(name="Arial", size=9, color="888888")
        for c in range(1, 4):
            ws_ids.cell(row=linha_atual, column=c).fill = PatternFill(
                "solid", fgColor=COR_CINZA_LINHA if linha_atual % 2 == 0 else COR_BRANCO)
            ws_ids.cell(row=linha_atual, column=c).border = borda_fina()
        ws_ids.row_dimensions[linha_atual].height = 18
        linha_atual += 1
    linha_atual += 1

ws_ids.column_dimensions["A"].width = 8
ws_ids.column_dimensions["B"].width = 40
ws_ids.column_dimensions["C"].width = 20

# ══════════════════════════════════════════════════════════════════════════════
# ABA 1 — UNIDADES DE MEDIDA
# ══════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("1️⃣ Unidades de Medida")
ws.sheet_properties.tabColor = "43A047"
ws.freeze_panes = "A5"
linha_header_principal(ws, 1, "1. UNIDADES DE MEDIDA", 3)
linha_legenda(ws, 2, 3)
cols = [("id\n🔵 AUTO", 10, COR_HEADER_MEDIO),
        ("nome\n🟡 OBRIGATÓRIO", 30, COR_HEADER_MEDIO),
        ("sigla\n🟡 OBRIGATÓRIO", 15, COR_HEADER_MEDIO)]
linha_header_col(ws, 3, cols)
celula(ws, 4, 1, "← deixe vazio", italico=True, cor_fundo=COR_INSTRUCAO, cor_texto="888888")
celula(ws, 4, 2, "Nome completo da unidade", italico=True, cor_fundo=COR_INSTRUCAO, cor_texto="888888")
celula(ws, 4, 3, "Abreviação", italico=True, cor_fundo=COR_INSTRUCAO, cor_texto="888888")
exemplos_unidade = [
    ("", "unidade", "un"),
    ("", "metro", "m"),
    ("", "metro quadrado", "m²"),
    ("", "quilograma", "kg"),
    ("", "litro", "L"),
    ("", "conjunto", "cj"),
]
for i, row in enumerate(exemplos_unidade, 5):
    cor = COR_EXEMPLO if i == 5 else (COR_CINZA_LINHA if i % 2 == 0 else COR_BRANCO)
    for j, v in enumerate(row, 1):
        celula(ws, i, j, v, cor_fundo=cor)
for i in range(len(exemplos_unidade) + 5, 30):
    for j in range(1, 4):
        celula(ws, i, j, "", cor_fundo=COR_CINZA_LINHA if i % 2 == 0 else COR_BRANCO)

# ══════════════════════════════════════════════════════════════════════════════
# ABA 2 — FABRICANTES
# ══════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("2️⃣ Fabricantes")
ws.sheet_properties.tabColor = "E53935"
ws.freeze_panes = "A5"
linha_header_principal(ws, 1, "2. FABRICANTES", 2)
linha_legenda(ws, 2, 2)
linha_header_col(ws, 3, [("id\n🔵 AUTO", 10, COR_HEADER_MEDIO),
                          ("nome\n🟡 OBRIGATÓRIO (único)", 35, COR_HEADER_MEDIO)])
celula(ws, 4, 1, "← deixe vazio", italico=True, cor_fundo=COR_INSTRUCAO, cor_texto="888888")
celula(ws, 4, 2, "Nome exato do fabricante", italico=True, cor_fundo=COR_INSTRUCAO, cor_texto="888888")
fabricantes = ["Tecumseh","Embraco","Bitzer","Danfoss","Parker","Elgin","Copeland","Heatcraft","Hussmann","Friga-Bohn","Genérico"]
for i, nome in enumerate(fabricantes, 5):
    cor = COR_EXEMPLO if i == 5 else (COR_CINZA_LINHA if i % 2 == 0 else COR_BRANCO)
    celula(ws, i, 1, "", cor_fundo=cor)
    celula(ws, i, 2, nome, cor_fundo=cor)
for i in range(len(fabricantes) + 5, 30):
    celula(ws, i, 1, "", cor_fundo=COR_CINZA_LINHA if i % 2 == 0 else COR_BRANCO)
    celula(ws, i, 2, "", cor_fundo=COR_CINZA_LINHA if i % 2 == 0 else COR_BRANCO)

# ══════════════════════════════════════════════════════════════════════════════
# ABA 3 — CATEGORIAS
# ══════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("3️⃣ Categorias")
ws.sheet_properties.tabColor = "FB8C00"
ws.freeze_panes = "A5"
linha_header_principal(ws, 1, "3. CATEGORIAS", 2)
linha_legenda(ws, 2, 2)
linha_header_col(ws, 3, [("id\n🔵 AUTO", 10, COR_HEADER_MEDIO),
                          ("nome\n🟡 OBRIGATÓRIO (único)", 40, COR_HEADER_MEDIO)])
celula(ws, 4, 1, "← deixe vazio", italico=True, cor_fundo=COR_INSTRUCAO, cor_texto="888888")
celula(ws, 4, 2, "Nome da categoria", italico=True, cor_fundo=COR_INSTRUCAO, cor_texto="888888")
categorias = ["Condensadora","Evaporadora","Compressor","Válvula de Expansão","Filtro Secador",
              "Visor de Líquido","Válvula Solenoide","Pressostato","Tubulação de Cobre",
              "Isolamento Térmico","Material Elétrico","Painel Frigorífico","Solda e Fluxo","Acessórios de Instalação"]
for i, nome in enumerate(categorias, 5):
    cor = COR_EXEMPLO if i == 5 else (COR_CINZA_LINHA if i % 2 == 0 else COR_BRANCO)
    celula(ws, i, 1, "", cor_fundo=cor)
    celula(ws, i, 2, nome, cor_fundo=cor)
for i in range(len(categorias) + 5, 35):
    celula(ws, i, 1, ""); celula(ws, i, 2, "")

# ══════════════════════════════════════════════════════════════════════════════
# ABA 4 — TIPOS DE PRODUTO
# ══════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("4️⃣ Tipos de Produto")
ws.sheet_properties.tabColor = "6D4C41"
ws.freeze_panes = "A5"
linha_header_principal(ws, 1, "4. TIPOS DE PRODUTO TÉRMICO", 2)
linha_legenda(ws, 2, 2)
linha_header_col(ws, 3, [("id\n🔵 AUTO", 10, COR_HEADER_MEDIO),
                          ("nome\n🟡 OBRIGATÓRIO (único)", 45, COR_HEADER_MEDIO)])
celula(ws, 4, 1, "← deixe vazio", italico=True, cor_fundo=COR_INSTRUCAO, cor_texto="888888")
celula(ws, 4, 2, "Nome do tipo/grupo de produto", italico=True, cor_fundo=COR_INSTRUCAO, cor_texto="888888")
tipos = ["Carnes e Aves","Laticínios","FLV (Frutas, Legumes e Verduras)","Pescados",
         "Frios e Embutidos","Sorvetes e Congelados","Bebidas","Padaria e Confeitaria","Geral / Industrial"]
for i, nome in enumerate(tipos, 5):
    cor = COR_EXEMPLO if i == 5 else (COR_CINZA_LINHA if i % 2 == 0 else COR_BRANCO)
    celula(ws, i, 1, "", cor_fundo=cor)
    celula(ws, i, 2, nome, cor_fundo=cor)
for i in range(len(tipos) + 5, 30):
    celula(ws, i, 1, ""); celula(ws, i, 2, "")

# ══════════════════════════════════════════════════════════════════════════════
# ABA 5 — PERFIS DE PRODUTO
# ══════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("5️⃣ Perfis de Produto")
ws.sheet_properties.tabColor = "D32F2F"
ws.freeze_panes = "A5"
linha_header_principal(ws, 1, "5. PERFIS DE PRODUTO TÉRMICO  (Valores ASHRAE / Dados Técnicos)", 12)
linha_legenda(ws, 2, 12)
cols5 = [
    ("id\n🔵 AUTO", 8, COR_HEADER_MEDIO),
    ("nome\n🟡 OBRIG.", 22, COR_HEADER_MEDIO),
    ("tipo_id\n🟣 FK", 10, "5C6BC0"),
    ("tipo_nome\n(referência)", 28, "78909C"),
    ("ponto_congelamento\n°C  🟡", 14, COR_HEADER_MEDIO),
    ("cp_acima\nkJ/kg°C 🟡", 12, COR_HEADER_MEDIO),
    ("calor_latente\nkJ/kg  🟡", 13, COR_HEADER_MEDIO),
    ("cp_abaixo\nkJ/kg°C 🟡", 12, COR_HEADER_MEDIO),
    ("taxa_respiracao\nW/ton 🟢", 13, "388E3C"),
    ("temp_conservacao\n°C  🟢", 13, "388E3C"),
    ("umidade_relativa\n%   🟢", 12, "388E3C"),
    ("teor_agua\n%   🟢", 10, "388E3C"),
]
linha_header_col(ws, 3, cols5)
instr5 = ["← vazio","Nome do produto","ID do tipo\n(ver aba IDs)","Só referência\nnão é importado",
          "Ex: -1.70","Ex: 3.52","Ex: 249.00","Ex: 1.76","FLV: preencher\noutros: vazio",
          "Ex: 2.0","Ex: 88.0","Ex: 74.0"]
for j, txt in enumerate(instr5, 1):
    celula(ws, 4, j, txt, italico=True, cor_fundo=COR_INSTRUCAO, cor_texto="888888", wrap=True)
ws.row_dimensions[4].height = 30

perfis = [
    ("","Carne Bovina",     1,"Carnes e Aves",  -1.70,3.52,249.0,1.76,None, 2.0,88.0,74.0),
    ("","Carne Suína",      1,"Carnes e Aves",  -2.00,3.44,243.0,1.72,None, 2.0,85.0,72.0),
    ("","Frango Inteiro",   1,"Carnes e Aves",  -2.80,3.31,246.0,1.55,None,-1.0,85.0,66.0),
    ("","Carne Congelada",  1,"Carnes e Aves",  -1.70,3.52,249.0,1.76,None,-18.,90.0,74.0),
    ("","Queijo Maturado",  2,"Laticínios",    -10.00,2.09, 84.0,1.26,None, 6.0,80.0,37.0),
    ("","Leite Pasteurizado",2,"Laticínios",    -0.60,3.93,268.0,1.93,None, 4.0,85.0,87.0),
    ("","Manteiga",         2,"Laticínios",    -2.30,2.05,113.0,1.26,None, 4.0,80.0,16.0),
    ("","Alface",           3,"FLV",           -0.20,3.96,256.0,1.97,62.0, 2.0,95.0,95.0),
    ("","Tomate Maduro",    3,"FLV",           -0.50,3.94,255.0,1.96,28.0,10.0,90.0,94.0),
    ("","Banana Madura",    3,"FLV",           -0.80,3.35,224.0,1.65,25.0,14.0,90.0,74.0),
    ("","Maçã",             3,"FLV",           -1.50,3.73,243.0,1.84,10.0, 2.0,92.0,84.0),
    ("","Batata",           3,"FLV",           -0.60,3.51,236.0,1.77, 9.0, 8.0,90.0,77.0),
    ("","Peixe Fresco",     4,"Pescados",      -2.20,3.76,258.0,1.86,None, 0.0,95.0,76.0),
    ("","Camarão Fresco",   4,"Pescados",      -2.30,3.60,260.0,1.77,None, 0.0,95.0,76.0),
    ("","Peixe Congelado",  4,"Pescados",      -2.20,3.76,258.0,1.86,None,-18.,90.0,76.0),
    ("","Presunto Cozido",  5,"Frios e Embutidos",-3.00,3.18,213.0,1.55,None,4.0,80.0,60.0),
    ("","Salsicha",         5,"Frios e Embutidos",-2.50,3.39,234.0,1.68,None,4.0,80.0,63.0),
    ("","Sorvete",          6,"Sorvetes",     -14.50,1.63,210.0,1.26,None,-18.,90.0,60.0),
    ("","Mix de Açaí",      6,"Sorvetes",     -10.00,2.10,200.0,1.30,None,-18.,90.0,60.0),
    ("","Cerveja Lata",     7,"Bebidas",       -2.20,4.02,268.0,1.97,None, 4.0,70.0,92.0),
    ("","Refrigerante",     7,"Bebidas",       -1.00,4.00,268.0,1.97,None, 4.0,70.0,90.0),
    ("","Vinho",            7,"Bebidas",       -4.20,3.94,264.0,1.95,None, 8.0,70.0,87.0),
    ("","Massa Fresca",     8,"Padaria",       -3.50,3.30,236.0,1.62,None, 4.0,85.0,65.0),
    ("","Chocolate",        8,"Padaria",      -17.60,1.47,105.0,1.05,None,15.0,50.0, 1.0),
    ("","Prod. Genérico +5",9,"Geral",          0.00,3.50,230.0,1.75,None, 5.0,80.0,70.0),
    ("","Prod. Genérico -18",9,"Geral",         -2.00,3.50,230.0,1.75,None,-18.,90.0,70.0),
]
for i, row in enumerate(perfis, 5):
    cor = COR_EXEMPLO if i == 5 else (COR_CINZA_LINHA if i % 2 == 0 else COR_BRANCO)
    for j, v in enumerate(row, 1):
        c = celula(ws, i, j, v, cor_fundo=cor)
        if j == 4:  # coluna referência
            c.font = Font(name="Arial", italic=True, size=9, color="888888")
for i in range(len(perfis) + 5, 80):
    for j in range(1, 13):
        celula(ws, i, j, "", cor_fundo=COR_CINZA_LINHA if i % 2 == 0 else COR_BRANCO)

# ══════════════════════════════════════════════════════════════════════════════
# ABA 6 — EQUIPAMENTOS
# ══════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("6️⃣ Equipamentos")
ws.sheet_properties.tabColor = "1565C0"
ws.freeze_panes = "A5"
linha_header_principal(ws, 1, "6. EQUIPAMENTOS (Condensadoras, Evaporadoras, Compressores)", 12)
linha_legenda(ws, 2, 12)
cols6 = [
    ("id\n🔵 AUTO", 8, COR_HEADER_MEDIO),
    ("modelo\n🟡 OBRIG.", 18, COR_HEADER_MEDIO),
    ("categoria_id\n🟣 FK", 12, "5C6BC0"),
    ("categoria\n(referência)", 18, "78909C"),
    ("fabricante_id\n🟣 FK", 12, "5C6BC0"),
    ("fabricante\n(referência)", 16, "78909C"),
    ("unidade_id\n🟣 FK", 10, "5C6BC0"),
    ("custo\nR$ 🟡", 12, COR_HEADER_MEDIO),
    ("qtde_vent.\n🟡", 10, COR_HEADER_MEDIO),
    ("diam_vent.\nmm 🟡", 10, COR_HEADER_MEDIO),
    ("vazao_ar\nm³/h 🟡", 10, COR_HEADER_MEDIO),
    ("flecha_ar\nm 🟡", 10, COR_HEADER_MEDIO),
]
linha_header_col(ws, 3, cols6)
instr6 = ["← vazio","Ex: CAJ2464Z","Ex: 1","Ex: Condensadora","Ex: 1","Ex: Tecumseh","1=unidade",
          "Ex: 1850.00","Ex: 1","Ex: 300","Ex: 1200","Ex: 4"]
for j, txt in enumerate(instr6, 1):
    celula(ws, 4, j, txt, italico=True, cor_fundo=COR_INSTRUCAO, cor_texto="888888")

equips = [
    # Condensadoras Tecumseh
    ("","CAJ2464Z",  1,"Condensadora", 1,"Tecumseh",1, 1850.00,1,300,1200,4),
    ("","CAJ4519Z",  1,"Condensadora", 1,"Tecumseh",1, 2300.00,1,350,1800,5),
    ("","CAJ9513Z",  1,"Condensadora", 1,"Tecumseh",1, 3200.00,1,400,3200,5),
    ("","TAJ4519Z",  1,"Condensadora", 1,"Tecumseh",1, 4500.00,2,400,4500,6),
    # Condensadoras Elgin
    ("","CHCB-5000", 1,"Condensadora", 6,"Elgin",   1, 2100.00,1,350,2000,5),
    ("","CHCB-8000", 1,"Condensadora", 6,"Elgin",   1, 3400.00,1,400,3800,6),
    ("","CHCB-12000",1,"Condensadora", 6,"Elgin",   1, 5200.00,2,450,5500,7),
    # Evaporadoras Tecumseh
    ("","SILP22",    2,"Evaporadora",  1,"Tecumseh",1,  980.00,1,300,1200,3),
    ("","SILP44",    2,"Evaporadora",  1,"Tecumseh",1, 1450.00,2,300,2400,3),
    ("","SILP66",    2,"Evaporadora",  1,"Tecumseh",1, 1950.00,2,350,3200,4),
    # Evaporadoras Elgin
    ("","EVAP-3000", 2,"Evaporadora",  6,"Elgin",   1,  850.00,1,300,1500,3),
    ("","EVAP-6000", 2,"Evaporadora",  6,"Elgin",   1, 1350.00,2,300,2800,3),
    ("","EVAP-10000",2,"Evaporadora",  6,"Elgin",   1, 2100.00,2,350,4500,4),
    # Compressores (sem ventilador)
    ("","EMI30HER",  3,"Compressor",   2,"Embraco", 1,  620.00,0,0,0,0),
    ("","NJ9232GK",  3,"Compressor",   2,"Embraco", 1,  980.00,0,0,0,0),
    ("","CAJ2464Z-C",3,"Compressor",   1,"Tecumseh",1,  750.00,0,0,0,0),
]
for i, row in enumerate(equips, 5):
    cor = COR_EXEMPLO if i == 5 else (COR_CINZA_LINHA if i % 2 == 0 else COR_BRANCO)
    for j, v in enumerate(row, 1):
        c = celula(ws, i, j, v, cor_fundo=cor)
        if j in (4, 6):
            c.font = Font(name="Arial", italic=True, size=9, color="888888")
for i in range(len(equips) + 5, 80):
    for j in range(1, 13):
        celula(ws, i, j, "", cor_fundo=COR_CINZA_LINHA if i % 2 == 0 else COR_BRANCO)

# ══════════════════════════════════════════════════════════════════════════════
# ABA 7 — PERFORMANCE EQUIPAMENTO
# ══════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("7️⃣ Performance Equip.")
ws.sheet_properties.tabColor = "00838F"
ws.freeze_panes = "A5"
linha_header_principal(ws, 1, "7. PERFORMANCE DOS EQUIPAMENTOS  ⚠️ Dados dos catálogos técnicos dos fabricantes", 10)
linha_legenda(ws, 2, 10)

# Nota importante
ws.merge_cells("A3:J3")
nota = ws.cell(row=3, column=1,
    value="⚠️  ATENÇÃO: A capacidade (kcal/h) DEVE ser retirada do catálogo técnico do fabricante para cada fluido e temperatura de evaporação. Valores incorretos resultam em seleção errada de equipamentos!")
nota.font = Font(name="Arial", bold=True, size=9, color="B71C1C")
nota.fill = PatternFill("solid", fgColor="FFEBEE")
nota.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws.row_dimensions[3].height = 30

cols7 = [
    ("id\n🔵 AUTO", 8, COR_HEADER_MEDIO),
    ("equipamento_id\n🟣 FK", 14, "5C6BC0"),
    ("modelo\n(referência)", 16, "78909C"),
    ("fluido\n🟡 OBRIG.", 12, COR_HEADER_MEDIO),
    ("temp_condensacao\n°C (padrão 45)", 14, COR_HEADER_MEDIO),
    ("temp_evaporacao\n°C 🟡", 14, COR_HEADER_MEDIO),
    ("delta_t\n°C (padrão 8)", 12, COR_HEADER_MEDIO),
    ("capacidade\nkcal/h 🟡", 14, COR_HEADER_MEDIO),
    ("consumo_w\nW 🟢", 12, "388E3C"),
    ("FONTE\n(anotação)", 25, "78909C"),
]
linha_header_col(ws, 4, cols7)

perfs_equip = [
    # CAJ2464Z id=1 R404A
    ("",1,"CAJ2464Z","R404A",45,-10,8,2100,680,"Catálogo Tecumseh 2024"),
    ("",1,"CAJ2464Z","R404A",45,-15,8,1700,620,"Catálogo Tecumseh 2024"),
    ("",1,"CAJ2464Z","R404A",45,-20,8,1350,560,"Catálogo Tecumseh 2024"),
    ("",1,"CAJ2464Z","R404A",45,-25,8,1050,500,"Catálogo Tecumseh 2024"),
    ("",1,"CAJ2464Z","R404A",45,-30,8, 820,450,"Catálogo Tecumseh 2024"),
    ("",1,"CAJ2464Z","R290", 45,-10,8,2300,650,"Catálogo Tecumseh 2024"),
    ("",1,"CAJ2464Z","R290", 45,-15,8,1900,590,"Catálogo Tecumseh 2024"),
    ("",1,"CAJ2464Z","R290", 45,-20,8,1500,530,"Catálogo Tecumseh 2024"),
    # CAJ4519Z id=2
    ("",2,"CAJ4519Z","R404A",45,-10,8,3800,1200,"Catálogo Tecumseh 2024"),
    ("",2,"CAJ4519Z","R404A",45,-15,8,3100,1100,"Catálogo Tecumseh 2024"),
    ("",2,"CAJ4519Z","R404A",45,-20,8,2500, 980,"Catálogo Tecumseh 2024"),
    ("",2,"CAJ4519Z","R404A",45,-25,8,1950, 860,"Catálogo Tecumseh 2024"),
    ("",2,"CAJ4519Z","R404A",45,-30,8,1500, 750,"Catálogo Tecumseh 2024"),
    # EVAP-3000 id=11
    ("",11,"EVAP-3000","R404A",45,-10,8,2200,0,"Catálogo Elgin 2024"),
    ("",11,"EVAP-3000","R404A",45,-15,8,1800,0,"Catálogo Elgin 2024"),
    ("",11,"EVAP-3000","R404A",45,-20,8,1450,0,"Catálogo Elgin 2024"),
    ("",11,"EVAP-3000","R404A",45,-25,8,1100,0,"Catálogo Elgin 2024"),
    ("",11,"EVAP-3000","R404A",45,-30,8, 850,0,"Catálogo Elgin 2024"),
]
for i, row in enumerate(perfs_equip, 5):
    cor = COR_EXEMPLO if i == 5 else (COR_CINZA_LINHA if i % 2 == 0 else COR_BRANCO)
    for j, v in enumerate(row, 1):
        c = celula(ws, i, j, v, cor_fundo=cor)
        if j == 3:
            c.font = Font(name="Arial", italic=True, size=9, color="888888")
        if j == 10:
            c.font = Font(name="Arial", italic=True, size=9, color="1565C0")
for i in range(len(perfs_equip) + 5, 200):
    for j in range(1, 11):
        celula(ws, i, j, "", cor_fundo=COR_CINZA_LINHA if i % 2 == 0 else COR_BRANCO)

# ══════════════════════════════════════════════════════════════════════════════
# ABA 8 — MATERIAIS
# ══════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("8️⃣ Materiais")
ws.sheet_properties.tabColor = "558B2F"
ws.freeze_panes = "A5"
linha_header_principal(ws, 1, "8. MATERIAIS (Tubulação, Isolamento, Elétrico, Painéis)", 11)
linha_legenda(ws, 2, 11)
cols8 = [
    ("id\n🔵 AUTO", 8, COR_HEADER_MEDIO),
    ("nome\n🟡 OBRIG.", 30, COR_HEADER_MEDIO),
    ("categoria_id\n🟣 FK", 12, "5C6BC0"),
    ("categoria\n(referência)", 22, "78909C"),
    ("fabricante_id\n🟣 FK 🟢", 13, "5C6BC0"),
    ("unidade_id\n🟣 FK", 10, "5C6BC0"),
    ("unidade\n(ref.)", 10, "78909C"),
    ("custo\nR$ 🟡", 12, COR_HEADER_MEDIO),
    ("diametro\nconexao 🟢", 12, "388E3C"),
    ("capacidade\nnominal 🟢", 12, "388E3C"),
    ("detalhes_tecnicos\nJSON 🟢", 28, "388E3C"),
]
linha_header_col(ws, 3, cols8)
instr8 = ["← vazio","Descrição completa","ID categoria","Só referência","ID fabricante\n(vazio=genérico)",
          "ID unidade","Só referência","Ex: 12.50","Ex: 3/8\"","0 se n/a","Ex: {\"esp_mm\":9}"]
for j, txt in enumerate(instr8, 1):
    celula(ws, 4, j, txt, italico=True, cor_fundo=COR_INSTRUCAO, cor_texto="888888", wrap=True)
ws.row_dimensions[4].height = 30

materiais = [
    ("","Tubo Cobre 1/4\"",  9,"Tubulação",None,2,"metro",  12.50,'1/4"', 0,'{"bitola":"1/4","esp_mm":0.8}'),
    ("","Tubo Cobre 3/8\"",  9,"Tubulação",None,2,"metro",  18.00,'3/8"', 0,'{"bitola":"3/8","esp_mm":0.8}'),
    ("","Tubo Cobre 1/2\"",  9,"Tubulação",None,2,"metro",  24.00,'1/2"', 0,'{"bitola":"1/2","esp_mm":0.9}'),
    ("","Tubo Cobre 5/8\"",  9,"Tubulação",None,2,"metro",  32.00,'5/8"', 0,'{"bitola":"5/8","esp_mm":0.9}'),
    ("","Tubo Cobre 3/4\"",  9,"Tubulação",None,2,"metro",  44.00,'3/4"', 0,'{"bitola":"3/4","esp_mm":1.0}'),
    ("","Tubo Cobre 7/8\"",  9,"Tubulação",None,2,"metro",  58.00,'7/8"', 0,'{"bitola":"7/8","esp_mm":1.0}'),
    ("","Tubo Cobre 1.1/8\"",9,"Tubulação",None,2,"metro",  88.00,'1.1/8"',0,'{"bitola":"1.1/8","esp_mm":1.1}'),
    ("","Isolamento 3/8\" 9mm",10,"Isolamento",None,2,"metro",10.00,'3/8"',0,'{"esp_mm":9}'),
    ("","Isolamento 1/2\" 13mm",10,"Isolamento",None,2,"metro",13.00,'1/2"',0,'{"esp_mm":13}'),
    ("","Isolamento 5/8\" 13mm",10,"Isolamento",None,2,"metro",16.00,'5/8"',0,'{"esp_mm":13}'),
    ("","Isolamento 3/4\" 19mm",10,"Isolamento",None,2,"metro",22.00,'3/4"',0,'{"esp_mm":19}'),
    ("","Painel PUR 75mm",  12,"Painel",None,3,"m²",   320.00,None,0,'{"esp_mm":75,"nucleo":"PUR"}'),
    ("","Painel PUR 100mm", 12,"Painel",None,3,"m²",   380.00,None,0,'{"esp_mm":100,"nucleo":"PUR"}'),
    ("","Painel PIR 100mm", 12,"Painel",None,3,"m²",   450.00,None,0,'{"esp_mm":100,"nucleo":"PIR"}'),
    ("","Solda Riacho 15%", 13,"Solda", None,4,"kg",    85.00,None,0,'{}'),
    ("","Fluxo Decapante",  13,"Solda", None,4,"kg",    32.00,None,0,'{}'),
    ("","Cabo PP 2x1.5mm",  11,"Elétrico",None,2,"metro",2.80,None,0,'{"secao":"2x1.5"}'),
    ("","Cabo PP 2x2.5mm",  11,"Elétrico",None,2,"metro",3.80,None,0,'{"secao":"2x2.5"}'),
    ("","Disjuntor 16A",    11,"Elétrico",None,1,"unidade",28.00,None,0,'{"amperes":16}'),
    ("","Disjuntor 25A",    11,"Elétrico",None,1,"unidade",35.00,None,0,'{"amperes":25}'),
]
for i, row in enumerate(materiais, 5):
    cor = COR_EXEMPLO if i == 5 else (COR_CINZA_LINHA if i % 2 == 0 else COR_BRANCO)
    for j, v in enumerate(row, 1):
        c = celula(ws, i, j, v, cor_fundo=cor)
        if j in (4, 7):
            c.font = Font(name="Arial", italic=True, size=9, color="888888")
for i in range(len(materiais) + 5, 100):
    for j in range(1, 12):
        celula(ws, i, j, "", cor_fundo=COR_CINZA_LINHA if i % 2 == 0 else COR_BRANCO)

# ══════════════════════════════════════════════════════════════════════════════
# ABA 9 — COMPONENTES
# ══════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("9️⃣ Componentes")
ws.sheet_properties.tabColor = "AD1457"
ws.freeze_panes = "A5"
linha_header_principal(ws, 1, "9. COMPONENTES TÉCNICOS (Válvulas, Filtros, Visores, Presostatos)", 11)
linha_legenda(ws, 2, 11)
cols9 = [
    ("id\n🔵 AUTO", 8, COR_HEADER_MEDIO),
    ("modelo\n🟡 OBRIG.", 18, COR_HEADER_MEDIO),
    ("categoria_id\n🟣 FK", 12, "5C6BC0"),
    ("categoria\n(referência)", 22, "78909C"),
    ("fabricante_id\n🟣 FK", 12, "5C6BC0"),
    ("fabricante\n(referência)", 14, "78909C"),
    ("cod_fabricante\n🟢", 16, "388E3C"),
    ("conexao_entrada\n🟡", 13, COR_HEADER_MEDIO),
    ("conexao_saida\n🟡", 13, COR_HEADER_MEDIO),
    ("capacidade\nkcal/h 🟡", 13, COR_HEADER_MEDIO),
    ("custo\nR$ 🟡", 12, COR_HEADER_MEDIO),
]
linha_header_col(ws, 3, cols9)

componentes = [
    ("","T2 R404A 1/4\"",  4,"VET",  4,"Danfoss","T2-R404A",'1/4"','1/4"', 2500,185.00),
    ("","T4 R404A 3/8\"",  4,"VET",  4,"Danfoss","T4-R404A",'3/8"','3/8"', 5000,220.00),
    ("","T8 R404A 1/2\"",  4,"VET",  4,"Danfoss","T8-R404A",'1/2"','1/2"',10000,280.00),
    ("","T2 R290 1/4\"",   4,"VET",  4,"Danfoss","T2-R290", '1/4"','1/4"', 2500,195.00),
    ("","T4 R290 3/8\"",   4,"VET",  4,"Danfoss","T4-R290", '3/8"','3/8"', 5000,235.00),
    ("","DML 032S",        5,"Filtro",4,"Danfoss","DML032S", '1/4"','1/4"', 3000, 42.00),
    ("","DML 053S",        5,"Filtro",4,"Danfoss","DML053S", '3/8"','3/8"', 6000, 58.00),
    ("","DML 083S",        5,"Filtro",4,"Danfoss","DML083S", '1/2"','1/2"',12000, 75.00),
    ("","SGP 1/4\"",       6,"Visor", 5,"Parker", "SGP14",   '1/4"','1/4"', 3000, 55.00),
    ("","SGP 3/8\"",       6,"Visor", 5,"Parker", "SGP38",   '3/8"','3/8"', 6000, 68.00),
    ("","SGP 1/2\"",       6,"Visor", 5,"Parker", "SGP12",   '1/2"','1/2"',12000, 85.00),
    ("","EVR 3 1/4\"",     7,"Solenóide",4,"Danfoss","EVR3-14",'1/4"','1/4"',3000,145.00),
    ("","EVR 6 3/8\"",     7,"Solenóide",4,"Danfoss","EVR6-38",'3/8"','3/8"',6000,175.00),
    ("","EVR10 1/2\"",     7,"Solenóide",4,"Danfoss","EVR10-12",'1/2"','1/2"',12000,215.00),
    ("","KP5 Alta Pressão",8,"Pressostato",4,"Danfoss","KP5",'1/4"','1/4"',15000,185.00),
    ("","KP1 Baixa Pressão",8,"Pressostato",4,"Danfoss","KP1",'1/4"','1/4"',15000,165.00),
]
for i, row in enumerate(componentes, 5):
    cor = COR_EXEMPLO if i == 5 else (COR_CINZA_LINHA if i % 2 == 0 else COR_BRANCO)
    for j, v in enumerate(row, 1):
        c = celula(ws, i, j, v, cor_fundo=cor)
        if j in (4, 6):
            c.font = Font(name="Arial", italic=True, size=9, color="888888")
for i in range(len(componentes) + 5, 80):
    for j in range(1, 12):
        celula(ws, i, j, "", cor_fundo=COR_CINZA_LINHA if i % 2 == 0 else COR_BRANCO)

# ══════════════════════════════════════════════════════════════════════════════
# ABA 10 — PERFORMANCE COMPONENTE
# ══════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("🔟 Performance Comp.")
ws.sheet_properties.tabColor = "F57F17"
ws.freeze_panes = "A5"
linha_header_principal(ws, 1, "10. PERFORMANCE DOS COMPONENTES", 8)
linha_legenda(ws, 2, 8)
cols10 = [
    ("id\n🔵 AUTO", 8, COR_HEADER_MEDIO),
    ("componente_id\n🟣 FK", 14, "5C6BC0"),
    ("componente\n(referência)", 22, "78909C"),
    ("fluido\n🟡", 12, COR_HEADER_MEDIO),
    ("temp_evaporacao\n°C 🟡", 14, COR_HEADER_MEDIO),
    ("temp_condensacao\n°C (def:45) 🟡", 14, COR_HEADER_MEDIO),
    ("capacidade_kcalh\n🟡", 14, COR_HEADER_MEDIO),
    ("capacidade_min\nkcal/h 🟢", 14, "388E3C"),
]
linha_header_col(ws, 3, cols10)

perfs_comp = [
    ("",1,"T2 R404A 1/4\"","R404A",-10,45,2500,200),
    ("",1,"T2 R404A 1/4\"","R404A",-15,45,2200,180),
    ("",1,"T2 R404A 1/4\"","R404A",-20,45,1900,160),
    ("",1,"T2 R404A 1/4\"","R404A",-25,45,1600,130),
    ("",2,"T4 R404A 3/8\"","R404A",-10,45,5500,400),
    ("",2,"T4 R404A 3/8\"","R404A",-15,45,4800,350),
    ("",2,"T4 R404A 3/8\"","R404A",-20,45,4100,300),
    ("",2,"T4 R404A 3/8\"","R404A",-25,45,3400,250),
    ("",3,"T8 R404A 1/2\"","R404A",-10,45,11000,800),
    ("",3,"T8 R404A 1/2\"","R404A",-15,45, 9500,700),
    ("",3,"T8 R404A 1/2\"","R404A",-20,45, 8000,600),
    ("",3,"T8 R404A 1/2\"","R404A",-25,45, 6500,500),
]
for i, row in enumerate(perfs_comp, 5):
    cor = COR_EXEMPLO if i == 5 else (COR_CINZA_LINHA if i % 2 == 0 else COR_BRANCO)
    for j, v in enumerate(row, 1):
        c = celula(ws, i, j, v, cor_fundo=cor)
        if j == 3:
            c.font = Font(name="Arial", italic=True, size=9, color="888888")
for i in range(len(perfs_comp) + 5, 150):
    for j in range(1, 9):
        celula(ws, i, j, "", cor_fundo=COR_CINZA_LINHA if i % 2 == 0 else COR_BRANCO)

# ── Salvar ──────────────────────────────────────────────────────────────────
saida = r"C:\Users\User\PycharmProjects\ProjetistaV2\scripts\cadastro_banco_projetista.xlsx"
wb.save(saida)
print(f"Planilha salva em: {saida}")

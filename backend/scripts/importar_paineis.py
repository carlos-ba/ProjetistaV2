"""Importa painéis frigoríficos de uma planilha Excel para o banco.

Uso:
    cd backend
    ..\.venv\Scripts\python.exe scripts\importar_paineis.py <caminho_da_planilha.xlsx>

Colunas esperadas (aba "Painéis", 1ª linha = cabeçalho):
    produto | fabricante | nucleo | espessura_mm | largura_mm |
    comprimento_max_m | auto_portancia_mm | peso_kg_m2 | u_global | custo

Comportamento:
    - Cria o fabricante se ainda não existir.
    - Upsert por chave única (fabricante + nucleo + espessura + largura):
      atualiza se já existe, insere se novo.
    - u_global é obrigatório; linhas sem u_global são puladas com aviso.
"""
import asyncio
import sys
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import select

# Permite rodar a partir de backend/ (adiciona o pacote app ao path)
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from app.database.session import SessionLocal  # noqa: E402
from app.models.catalogo import Fabricante  # noqa: E402
from app.models.painel import PainelFrigorifico  # noqa: E402


def _num(v):
    if v is None or v == "":
        return None
    try:
        return Decimal(str(v))
    except (InvalidOperation, ValueError):
        return None


def _int(v):
    n = _num(v)
    return int(n) if n is not None else None


async def importar(caminho: str):
    wb = load_workbook(caminho, data_only=True)
    ws = wb["Painéis"] if "Painéis" in wb.sheetnames else wb.active

    # Mapeia cabeçalho → índice de coluna
    header = [str(c.value).strip() if c.value else "" for c in ws[1]]
    idx = {nome: i for i, nome in enumerate(header)}
    obrig = ["produto", "fabricante", "nucleo", "espessura_mm", "largura_mm", "u_global"]
    faltando = [c for c in obrig if c not in idx]
    if faltando:
        print(f"ERRO: colunas obrigatórias ausentes no cabeçalho: {faltando}")
        return

    def cel(row, nome):
        return row[idx[nome]] if nome in idx else None

    inseridos = atualizados = pulados = 0

    async with SessionLocal() as db:
        # cache de fabricantes por nome
        result = await db.execute(select(Fabricante))
        fab_por_nome = {f.nome.strip().lower(): f for f in result.scalars().all()}

        for row in ws.iter_rows(min_row=2, values_only=True):
            produto = cel(row, "produto")
            fab_nome = cel(row, "fabricante")
            nucleo = cel(row, "nucleo")
            espessura = _int(cel(row, "espessura_mm"))
            largura = _int(cel(row, "largura_mm"))
            u_global = _num(cel(row, "u_global"))

            # linha vazia
            if not produto and not fab_nome:
                continue
            # validação mínima
            if not (produto and fab_nome and nucleo and espessura and largura):
                pulados += 1
                continue
            if u_global is None:
                print(f"PULADA (sem u_global): {fab_nome} {nucleo} {espessura}mm {largura}mm")
                pulados += 1
                continue

            # fabricante (get or create)
            chave = str(fab_nome).strip().lower()
            fab = fab_por_nome.get(chave)
            if not fab:
                fab = Fabricante(nome=str(fab_nome).strip())
                db.add(fab)
                await db.flush()
                fab_por_nome[chave] = fab

            # upsert do painel pela chave única
            existente = await db.execute(
                select(PainelFrigorifico).where(
                    PainelFrigorifico.fabricante_id == fab.id,
                    PainelFrigorifico.nucleo == str(nucleo).strip(),
                    PainelFrigorifico.espessura_mm == espessura,
                    PainelFrigorifico.largura_mm == largura,
                )
            )
            p = existente.scalar_one_or_none()
            campos = dict(
                produto=str(produto).strip(),
                comprimento_max_m=_num(cel(row, "comprimento_max_m")),
                auto_portancia_mm=_int(cel(row, "auto_portancia_mm")),
                peso_kg_m2=_num(cel(row, "peso_kg_m2")),
                u_global=u_global,
                custo=_num(cel(row, "custo")) or Decimal("0"),
            )
            if p:
                for k, v in campos.items():
                    setattr(p, k, v)
                atualizados += 1
            else:
                db.add(PainelFrigorifico(
                    fabricante_id=fab.id,
                    nucleo=str(nucleo).strip(),
                    espessura_mm=espessura,
                    largura_mm=largura,
                    **campos,
                ))
                inseridos += 1

        await db.commit()

    print(f"OK — inseridos: {inseridos} | atualizados: {atualizados} | pulados: {pulados}")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Uso: python scripts/importar_paineis.py <planilha.xlsx>")
        sys.exit(1)
    # Windows: psycopg async exige SelectorEventLoop (não o ProactorEventLoop padrão)
    if sys.platform == "win32":
        asyncio.set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())
    asyncio.run(importar(sys.argv[1]))

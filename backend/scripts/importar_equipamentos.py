"""Importa equipamentos (UC ou Evaporadoras) + performance de uma planilha Excel.

Uso:
    cd backend
    ..\.venv\Scripts\python.exe scripts\importar_equipamentos.py <planilha.xlsx> "Unidade Condensadora"
    ..\.venv\Scripts\python.exe scripts\importar_equipamentos.py <planilha.xlsx> "Evaporadora"

Planilha (formato dos templates gerados):
    Aba "Equipamentos": modelo | fabricante | qtde_ventiladores | diametro_ventilador_mm |
        vazao_ar_m3h | flecha_ar_m | volume_interno_kg | conexao_liquido | conexao_succao | custo
    Aba "Performance":  modelo | fluido | T_ambiente_C | temp_evaporacao_C | delta_t |
        capacidade_kcalh | consumo_kw

Comportamento:
    - PULA as linhas cinza/itálicas (referência Elgin dos templates) — só importa as linhas novas.
    - Cria fabricante automaticamente se não existir.
    - Equipamento: upsert por (modelo + fabricante + categoria).
    - Performance: upsert pela chave única (equipamento + fluido + T.Amb + T.Evap + delta_t).
    - Linhas de performance cujo modelo não está na aba Equipamentos são buscadas no banco;
      se o modelo não existir em lugar nenhum, a linha é pulada com aviso.
"""
import asyncio
import sys
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import select

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from app.database.session import SessionLocal  # noqa: E402
from app.models.catalogo import Categoria, Fabricante, UnidadeMedida  # noqa: E402
from app.models.equipamento import Equipamento, PerformanceEquipamento  # noqa: E402

REF_COLOR = "808080"  # cinza das linhas de referência dos templates


def _num(v):
    if v is None or v == "":
        return None
    try:
        return Decimal(str(v))
    except (InvalidOperation, ValueError):
        return None


def _int(v):
    n = _num(v)
    return int(n) if n is not None else None


def _eh_referencia(ws, row_idx):
    """Linha de referência dos templates: fonte itálica/cinza na 1ª célula."""
    c = ws.cell(row=row_idx, column=1)
    f = c.font
    if f.italic:
        return True
    cor = getattr(f.color, "rgb", None)
    return bool(cor and str(cor).endswith(REF_COLOR))


async def importar(caminho: str, categoria_nome: str):
    wb = load_workbook(caminho)  # sem data_only para preservar formatação
    ws_eq = wb["Equipamentos"]
    ws_pf = wb["Performance"]

    async with SessionLocal() as db:
        # categoria
        cat = (await db.execute(
            select(Categoria).where(Categoria.nome == categoria_nome)
        )).scalar_one_or_none()
        if not cat:
            print(f"ERRO: categoria '{categoria_nome}' não existe no banco.")
            return

        # unidade de medida padrão ('un' — cria se necessário)
        um = (await db.execute(
            select(UnidadeMedida).where(UnidadeMedida.sigla == "un")
        )).scalars().first()
        if not um:
            um = (await db.execute(select(UnidadeMedida))).scalars().first()
        if not um:
            um = UnidadeMedida(nome="Unidade", sigla="un")
            db.add(um)
            await db.flush()

        # fabricantes em cache
        fab_por_nome = {
            f.nome.strip().lower(): f
            for f in (await db.execute(select(Fabricante))).scalars().all()
        }

        # ── Aba Equipamentos ──────────────────────────────────────────────
        eq_ins = eq_upd = eq_ref = 0
        equip_por_modelo: dict[str, Equipamento] = {}
        for i, row in enumerate(ws_eq.iter_rows(min_row=2, values_only=True), start=2):
            modelo, fab_nome = row[0], row[1]
            if not modelo:
                continue
            if _eh_referencia(ws_eq, i):
                eq_ref += 1
                continue
            modelo = str(modelo).strip()
            chave_fab = str(fab_nome or "").strip().lower()
            if not chave_fab:
                print(f"PULADA (sem fabricante): {modelo}")
                continue
            fab = fab_por_nome.get(chave_fab)
            if not fab:
                fab = Fabricante(nome=str(fab_nome).strip())
                db.add(fab)
                await db.flush()
                fab_por_nome[chave_fab] = fab

            existente = (await db.execute(
                select(Equipamento).where(
                    Equipamento.modelo == modelo,
                    Equipamento.fabricante_id == fab.id,
                    Equipamento.categoria_id == cat.id,
                )
            )).scalar_one_or_none()
            campos = dict(
                qtde_ventiladores=_int(row[2]) or 0,
                diametro_ventilador_mm=_int(row[3]) or 0,
                vazao_ar_m3h=_int(row[4]) or 0,
                flecha_ar_m=_int(row[5]) or 0,
                volume_interno_kg=_num(row[6]),
                conexao_liquido=(str(row[7]).strip() if row[7] else None),
                conexao_succao=(str(row[8]).strip() if row[8] else None),
                custo=_num(row[9]) or Decimal("0"),
            )
            if existente:
                for k, v in campos.items():
                    setattr(existente, k, v)
                equip_por_modelo[modelo] = existente
                eq_upd += 1
            else:
                novo = Equipamento(
                    modelo=modelo, fabricante_id=fab.id,
                    categoria_id=cat.id, unidade_medida_id=um.id, **campos,
                )
                db.add(novo)
                await db.flush()
                equip_por_modelo[modelo] = novo
                eq_ins += 1

        # ── Aba Performance ───────────────────────────────────────────────
        pf_ins = pf_upd = pf_ref = pf_pul = 0
        for i, row in enumerate(ws_pf.iter_rows(min_row=2, values_only=True), start=2):
            modelo = row[0]
            if not modelo:
                continue
            if _eh_referencia(ws_pf, i):
                pf_ref += 1
                continue
            modelo = str(modelo).strip()
            fluido = str(row[1] or "").strip()
            t_amb = _int(row[2])
            t_evap = _int(row[3])
            delta_t = _num(row[4]) or Decimal("0")
            capacidade = _int(row[5])
            consumo = _num(row[6])
            if not (fluido and t_amb is not None and t_evap is not None and capacidade):
                print(f"PULADA (dados incompletos) linha {i}: {row}")
                pf_pul += 1
                continue

            eq = equip_por_modelo.get(modelo)
            if not eq:
                eq = (await db.execute(
                    select(Equipamento).where(
                        Equipamento.modelo == modelo,
                        Equipamento.categoria_id == cat.id,
                    )
                )).scalars().first()
                if not eq:
                    print(f"PULADA (modelo '{modelo}' não encontrado) linha {i}")
                    pf_pul += 1
                    continue
                equip_por_modelo[modelo] = eq

            existente = (await db.execute(
                select(PerformanceEquipamento).where(
                    PerformanceEquipamento.equipamento_id == eq.id,
                    PerformanceEquipamento.fluido == fluido,
                    PerformanceEquipamento.temp_ambiente == t_amb,
                    PerformanceEquipamento.temp_evaporacao == t_evap,
                    PerformanceEquipamento.delta_t == delta_t,
                )
            )).scalar_one_or_none()
            if existente:
                existente.capacidade = capacidade
                existente.consumo_kw = consumo
                pf_upd += 1
            else:
                db.add(PerformanceEquipamento(
                    equipamento_id=eq.id, fluido=fluido, temp_ambiente=t_amb,
                    temp_evaporacao=t_evap, delta_t=delta_t,
                    capacidade=capacidade, consumo_kw=consumo,
                ))
                pf_ins += 1

        await db.commit()

    print(f"EQUIPAMENTOS — inseridos: {eq_ins} | atualizados: {eq_upd} | referência (puladas): {eq_ref}")
    print(f"PERFORMANCE  — inseridas: {pf_ins} | atualizadas: {pf_upd} | referência: {pf_ref} | puladas: {pf_pul}")


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print('Uso: python scripts/importar_equipamentos.py <planilha.xlsx> "Unidade Condensadora"|"Evaporadora"')
        sys.exit(1)
    if sys.platform == "win32":
        asyncio.set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())
    asyncio.run(importar(sys.argv[1], sys.argv[2]))

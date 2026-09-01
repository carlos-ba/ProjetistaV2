"""Importa perfis metálicos de uma planilha Excel (catálogo do fabricante) para o banco.

Uso:
    cd backend
    ..\\.venv\\Scripts\\python.exe scripts\\importar_perfis_metalicos.py <planilha.xlsx> "<Nome do Fabricante>"

Colunas esperadas (1ª aba, 1ª linha = cabeçalho):
    CÓD.PRODUTO | DESCRIÇÃO DOS MATERIAIS

A descrição segue o padrão MBP Isoblock — "P <TIPO> ... <medidas separadas por x> <acabamento>":
    P ANG INT / P ANG EXT  -> Ângulo Interno / Externo (aba1 x aba2 x comprimento)
    P LISO                 -> Liso            (largura x comprimento)
    P U                    -> U               (aba1 x alma x aba2 x comprimento)
    P Z                    -> Z               (aba1 x aba2 x aba3 x comprimento)
O sufixo de acabamento (V10/V11/S-V) é ignorado, de propósito (confirmado com o usuário).
Linhas que não casam com nenhum tipo conhecido (ex: "PERFIL ESPECIAL") são puladas com aviso —
são itens avulsos fora do padrão dimensional, não entram neste catálogo.

Comportamento:
    - Cria o fabricante se ainda não existir.
    - Upsert por chave única (fabricante_id + codigo_fabricante): atualiza se já existe, insere se novo.
"""
import asyncio
import re
import sys
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import select

# Permite rodar a partir de backend/ (adiciona o pacote app ao path)
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from app.database.session import SessionLocal  # noqa: E402
from app.models.catalogo import Fabricante  # noqa: E402
from app.models.perfil_metalico import PerfilMetalico  # noqa: E402


_DIM_RE = re.compile(r"(\d+(?:x\d+){1,3})", re.IGNORECASE)

# (prefixo normalizado, tipo, qtde de medidas antes do comprimento)
_TIPOS = [
    ("P ANG INT", "Ângulo Interno", 2),
    ("P ANG EXT", "Ângulo Externo", 2),
    ("P LISO", "Liso", 1),
    ("P U", "U", 3),
    ("P Z", "Z", 3),
]


def _parse_linha(descricao: str):
    """Extrai (tipo, medida_1, medida_2, medida_3, comprimento) da descrição bruta,
    ou None se a linha não bate com nenhum tipo/padrão dimensional conhecido."""
    # Normaliza espaços múltiplos (ex: "P  LISO" com 2 espaços, ruído do fornecedor)
    desc = re.sub(r"\s+", " ", (descricao or "").strip())
    tipo_info = next((t for t in _TIPOS if desc.upper().startswith(t[0])), None)
    if not tipo_info:
        return None
    _, tipo, n_medidas = tipo_info

    m = _DIM_RE.search(desc)
    if not m:
        return None
    numeros = [int(x) for x in m.group(1).split("x")]
    if len(numeros) != n_medidas + 1:  # +1 é sempre o comprimento
        return None

    *medidas, comprimento = numeros
    medida_1 = medidas[0]
    medida_2 = medidas[1] if len(medidas) > 1 else None
    medida_3 = medidas[2] if len(medidas) > 2 else None
    return tipo, medida_1, medida_2, medida_3, comprimento


async def importar(caminho: str, fabricante_nome: str):
    wb = load_workbook(caminho, data_only=True)
    ws = wb.active

    header = [str(c.value).strip() if c.value else "" for c in ws[1]]
    idx = {nome: i for i, nome in enumerate(header)}
    obrig = ["CÓD.PRODUTO", "DESCRIÇÃO DOS MATERIAIS"]
    faltando = [c for c in obrig if c not in idx]
    if faltando:
        print(f"ERRO: colunas obrigatórias ausentes no cabeçalho: {faltando}")
        return

    def cel(row, nome):
        return row[idx[nome]] if nome in idx else None

    inseridos = atualizados = pulados = 0

    async with SessionLocal() as db:
        result = await db.execute(select(Fabricante).where(Fabricante.nome == fabricante_nome))
        fab = result.scalar_one_or_none()
        if not fab:
            fab = Fabricante(nome=fabricante_nome)
            db.add(fab)
            await db.flush()

        for row in ws.iter_rows(min_row=2, values_only=True):
            codigo = cel(row, "CÓD.PRODUTO")
            descricao = cel(row, "DESCRIÇÃO DOS MATERIAIS")
            if not codigo and not descricao:
                continue
            codigo = str(codigo).strip()
            descricao = str(descricao).strip()

            parsed = _parse_linha(descricao)
            if not parsed:
                print(f"PULADA (fora do padrão dimensional): {codigo} — {descricao}")
                pulados += 1
                continue
            tipo, m1, m2, m3, compr = parsed

            existente = await db.execute(
                select(PerfilMetalico).where(
                    PerfilMetalico.fabricante_id == fab.id,
                    PerfilMetalico.codigo_fabricante == codigo,
                )
            )
            p = existente.scalar_one_or_none()
            campos = dict(
                tipo=tipo, medida_1_mm=m1, medida_2_mm=m2, medida_3_mm=m3,
                comprimento_mm=compr, descricao_original=descricao,
            )
            if p:
                for k, v in campos.items():
                    setattr(p, k, v)
                atualizados += 1
            else:
                db.add(PerfilMetalico(fabricante_id=fab.id, codigo_fabricante=codigo, **campos))
                inseridos += 1

        await db.commit()

    print(f"OK — inseridos: {inseridos} | atualizados: {atualizados} | pulados: {pulados}")


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print('Uso: python scripts/importar_perfis_metalicos.py <planilha.xlsx> "<Nome do Fabricante>"')
        sys.exit(1)
    # Windows: psycopg async exige SelectorEventLoop (não o ProactorEventLoop padrão)
    if sys.platform == "win32":
        asyncio.set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())
    asyncio.run(importar(sys.argv[1], sys.argv[2]))

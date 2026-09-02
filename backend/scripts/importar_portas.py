"""Importa portas frigoríficas de uma planilha Excel para o banco.

Uso:
    cd backend
    ..\.venv\Scripts\python.exe scripts\importar_portas.py <caminho_da_planilha.xlsx>

Colunas esperadas (aba "Portas", 1ª linha = cabeçalho):
    descricao | fabricante | tipo | largura_mm | altura_mm | espessura_mm |
    classificacao | abertura | batente | soleira | custo | observacao

    tipo:          giratoria | deslizante | rapida
    classificacao: resfriados | congelada | ultra-congelada
    abertura:      direita | esquerda | ambas | automatica
    batente:       3B | 4B (texto livre — outros valores também aceitos)
    soleira:       TRUE/FALSE (aceita também 1/0, sim/não)

Comportamento:
    - Cria o fabricante se ainda não existir.
    - Upsert por chave única (fabricante + tipo + classificacao + largura +
      altura + espessura + batente + abertura): atualiza se já existe,
      insere se novo. Não há UniqueConstraint no banco — a checagem é feita
      aqui, mesmo padrão do importar_paineis.py.
    - descricao, tipo, largura_mm, altura_mm, espessura_mm e classificacao
      são obrigatórios; linha sem algum deles é pulada com aviso.
"""
import asyncio
import sys
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import select

# Permite rodar a partir de backend/ (adiciona o pacote app ao path)
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from app.database.session import SessionLocal  # noqa: E402
from app.models.catalogo import Fabricante  # noqa: E402
from app.models.porta import PortaFrigoriifica  # noqa: E402


def _num(v):
    if v is None or v == "":
        return None
    try:
        return Decimal(str(v))
    except (InvalidOperation, ValueError):
        return None


def _int(v):
    n = _num(v)
    return int(n) if n is not None else None


def _bool(v):
    if isinstance(v, bool):
        return v
    if v is None or v == "":
        return False
    return str(v).strip().lower() in ("true", "1", "sim", "verdadeiro", "x")


async def importar(caminho: str):
    wb = load_workbook(caminho, data_only=True)
    ws = wb["Portas"] if "Portas" in wb.sheetnames else wb.active

    header = [str(c.value).strip() if c.value else "" for c in ws[1]]
    idx = {nome: i for i, nome in enumerate(header)}
    obrig = ["descricao", "fabricante", "tipo", "largura_mm", "altura_mm", "espessura_mm", "classificacao"]
    faltando = [c for c in obrig if c not in idx]
    if faltando:
        print(f"ERRO: colunas obrigatórias ausentes no cabeçalho: {faltando}")
        return

    def cel(row, nome):
        return row[idx[nome]] if nome in idx else None

    inseridos = atualizados = pulados = 0

    async with SessionLocal() as db:
        # cache de fabricantes por nome
        result = await db.execute(select(Fabricante))
        fab_por_nome = {f.nome.strip().lower(): f for f in result.scalars().all()}

        for row in ws.iter_rows(min_row=2, values_only=True):
            descricao = cel(row, "descricao")
            fab_nome = cel(row, "fabricante")
            tipo = cel(row, "tipo")
            largura = _int(cel(row, "largura_mm"))
            altura = _int(cel(row, "altura_mm"))
            espessura = _int(cel(row, "espessura_mm"))
            classificacao = cel(row, "classificacao")
            abertura = cel(row, "abertura")
            batente = cel(row, "batente")

            # linha vazia
            if not descricao and not fab_nome:
                continue
            if not (descricao and fab_nome and tipo and largura and altura and espessura and classificacao):
                print(f"PULADA (faltam campos obrigatórios): {descricao or fab_nome}")
                pulados += 1
                continue

            # fabricante (get or create)
            chave = str(fab_nome).strip().lower()
            fab = fab_por_nome.get(chave)
            if not fab:
                fab = Fabricante(nome=str(fab_nome).strip())
                db.add(fab)
                await db.flush()
                fab_por_nome[chave] = fab

            abertura_v = str(abertura).strip() if abertura else None
            batente_v = str(batente).strip() if batente else None

            # upsert da porta pela chave única (checagem em código — não há
            # UniqueConstraint no banco pra essa tabela)
            existente = await db.execute(
                select(PortaFrigoriifica).where(
                    PortaFrigoriifica.fabricante_id == fab.id,
                    PortaFrigoriifica.tipo == str(tipo).strip(),
                    PortaFrigoriifica.classificacao == str(classificacao).strip(),
                    PortaFrigoriifica.largura_mm == largura,
                    PortaFrigoriifica.altura_mm == altura,
                    PortaFrigoriifica.espessura_mm == espessura,
                    PortaFrigoriifica.batente == batente_v,
                    PortaFrigoriifica.abertura == abertura_v,
                )
            )
            p = existente.scalar_one_or_none()
            campos = dict(
                descricao=str(descricao).strip(),
                soleira=_bool(cel(row, "soleira")),
                custo=_num(cel(row, "custo")) or Decimal("0"),
                observacao=(str(cel(row, "observacao")).strip() if cel(row, "observacao") else None),
            )
            if p:
                for k, v in campos.items():
                    setattr(p, k, v)
                atualizados += 1
            else:
                db.add(PortaFrigoriifica(
                    fabricante_id=fab.id,
                    tipo=str(tipo).strip(),
                    classificacao=str(classificacao).strip(),
                    largura_mm=largura,
                    altura_mm=altura,
                    espessura_mm=espessura,
                    batente=batente_v,
                    abertura=abertura_v,
                    **campos,
                ))
                inseridos += 1

        await db.commit()

    print(f"OK — inseridos: {inseridos} | atualizados: {atualizados} | pulados: {pulados}")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Uso: python scripts/importar_portas.py <planilha.xlsx>")
        sys.exit(1)
    # Windows: psycopg async exige SelectorEventLoop (não o ProactorEventLoop padrão)
    if sys.platform == "win32":
        asyncio.set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())
    asyncio.run(importar(sys.argv[1]))

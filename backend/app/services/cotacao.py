"""
Gera planilha Excel de cotação para fornecedor.

A planilha é enviada ao fornecedor com a lista de materiais e equipamentos.
O fornecedor preenche a coluna de preço unitário e devolve.
Na próxima fase, a planilha respondida pode ser reimportada para atualizar o banco.

Estrutura da planilha:
  Coluna A: Código interno (id do banco — referência para reimportação)
  Coluna B: Tipo (Equipamento / Material / Componente)
  Coluna C: Descrição do item
  Coluna D: Fabricante / Detalhe
  Coluna E: Quantidade
  Coluna F: Unidade
  Coluna G: Preço Unitário (R$)  ← fornecedor preenche
  Coluna H: Total                ← fórmula automática E×G
  Coluna I: Observações do fornecedor
"""
from __future__ import annotations
from io import BytesIO
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
)
from openpyxl.utils import get_column_letter

# ── Cores ─────────────────────────────────────────────────────────────────
COR_HEADER      = "1E3A5F"   # azul escuro
COR_EQUIP       = "EBF5FB"   # azul claro — equipamentos
COR_MATERIAL    = "FDFEFE"   # branco/cinza claro — materiais
COR_PREENCHER   = "FFFACD"   # amarelo suave — coluna a preencher
COR_TOTAL       = "EAF2FF"   # azul total
COR_RODAPE      = "F4F6F6"   # cinza claro

BORDA_FINA = Border(
    left  =Side(style='thin', color='CCCCCC'),
    right =Side(style='thin', color='CCCCCC'),
    top   =Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC'),
)

BORDA_HEADER = Border(
    left  =Side(style='medium', color='1E3A5F'),
    right =Side(style='medium', color='1E3A5F'),
    top   =Side(style='medium', color='1E3A5F'),
    bottom=Side(style='medium', color='1E3A5F'),
)


def _celula(ws, linha, col, valor, negrito=False, cor_fundo=None,
            cor_texto="000000", alinhamento="left", tamanho=10,
            borda=True, italico=False):
    c = ws.cell(row=linha, column=col, value=valor)
    c.font = Font(name="Arial", bold=negrito, italic=italico,
                  color=cor_texto, size=tamanho)
    if cor_fundo:
        c.fill = PatternFill("solid", fgColor=cor_fundo)
    c.alignment = Alignment(horizontal=alinhamento, vertical="center",
                             wrap_text=False)
    if borda:
        c.border = BORDA_FINA
    return c


def gerar_planilha_cotacao(
    itens: list[dict],          # lista de {id, tipo, item, detalhe, quantidade, unidade}
    nome_projeto: str = "Projeto",
    nome_cliente: str = "",
) -> bytes:
    """
    Retorna bytes do arquivo .xlsx pronto para download.

    Cada item deve ter:
      id         → código interno (pode ser None para itens do gabinete)
      tipo       → 'Equipamento' | 'Material' | 'Componente'
      item       → descrição do item
      detalhe    → fabricante, modelo, especificação técnica
      quantidade → float
      unidade    → string (un, m, m², kg...)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Cotação"
    ws.sheet_view.showGridLines = False

    # ── Cabeçalho do documento ────────────────────────────────────────────
    ws.merge_cells("A1:I1")
    c = ws.cell(row=1, column=1, value="PLANILHA DE COTAÇÃO — MATERIAIS E EQUIPAMENTOS")
    c.font      = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    c.fill      = PatternFill("solid", fgColor=COR_HEADER)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:I2")
    info = f"Projeto: {nome_projeto}"
    if nome_cliente:
        info += f" | Cliente: {nome_cliente}"
    info += f" | Data: {datetime.now().strftime('%d/%m/%Y')}"
    c2 = ws.cell(row=2, column=1, value=info)
    c2.font      = Font(name="Arial", size=9, color="555555", italic=True)
    c2.fill      = PatternFill("solid", fgColor="EBF2FF")
    c2.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    ws.merge_cells("A3:I3")
    inst = ws.cell(row=3, column=1,
        value="INSTRUÇÕES AO FORNECEDOR: Preencha apenas a coluna 'Preço Unitário (R$)' destacada em amarelo. Não altere as demais colunas.")
    inst.font      = Font(name="Arial", size=8, color="8B4513", bold=True)
    inst.fill      = PatternFill("solid", fgColor="FFFACD")
    inst.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[3].height = 16

    # ── Cabeçalho das colunas ─────────────────────────────────────────────
    COLUNAS = [
        ("Código",          8),
        ("Tipo",           14),
        ("Descrição",      42),
        ("Detalhe / Esp. Técnica", 30),
        ("Qtde",            8),
        ("Un",              6),
        ("Preço Unit. (R$)", 16),
        ("Total (R$)",      14),
        ("Obs. Fornecedor", 22),
    ]

    for col_idx, (titulo, largura) in enumerate(COLUNAS, 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = largura

        # Destaque especial na coluna G (Preço Unit.)
        cor = COR_PREENCHER if col_idx == 7 else COR_HEADER
        txt = "000000" if col_idx == 7 else "FFFFFF"

        c = ws.cell(row=4, column=col_idx, value=titulo)
        c.font      = Font(name="Arial", bold=True, size=9,
                           color=txt)
        c.fill      = PatternFill("solid", fgColor=cor)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = BORDA_HEADER

    ws.row_dimensions[4].height = 22

    # ── Linhas de itens ───────────────────────────────────────────────────
    lin = 5
    for item in itens:
        cor_linha = COR_EQUIP if item.get("tipo") == "Equipamento" else COR_MATERIAL

        _celula(ws, lin, 1, str(item.get("id") or "-"),      cor_fundo=cor_linha, alinhamento="center", tamanho=8, cor_texto="888888")
        _celula(ws, lin, 2, item.get("tipo", ""),             cor_fundo=cor_linha, tamanho=9)
        _celula(ws, lin, 3, item.get("item", ""),             cor_fundo=cor_linha, negrito=True, tamanho=9)
        _celula(ws, lin, 4, item.get("detalhe", ""),          cor_fundo=cor_linha, tamanho=8, cor_texto="555555")
        _celula(ws, lin, 5, item.get("quantidade", 1),        cor_fundo=cor_linha, alinhamento="center", tamanho=9)
        _celula(ws, lin, 6, item.get("unidade", "un"),        cor_fundo=cor_linha, alinhamento="center", tamanho=9)

        # Coluna G: preço unitário — a ser preenchida (amarelo, sem valor)
        cg = ws.cell(row=lin, column=7, value=None)
        cg.fill      = PatternFill("solid", fgColor=COR_PREENCHER)
        cg.border    = BORDA_FINA
        cg.alignment = Alignment(horizontal="right", vertical="center")
        cg.number_format = '#,##0.00'

        # Coluna H: total — fórmula automática
        col_e = get_column_letter(5)
        col_g = get_column_letter(7)
        ch = ws.cell(row=lin, column=8, value=f"=IF({col_g}{lin}=\"\",\"\",{col_e}{lin}*{col_g}{lin})")
        ch.fill         = PatternFill("solid", fgColor=COR_TOTAL)
        ch.border       = BORDA_FINA
        ch.alignment    = Alignment(horizontal="right", vertical="center")
        ch.number_format = '#,##0.00'
        ch.font         = Font(name="Arial", bold=True, size=9)

        # Coluna I: observações
        ci = ws.cell(row=lin, column=9, value=None)
        ci.fill   = PatternFill("solid", fgColor=COR_MATERIAL)
        ci.border = BORDA_FINA

        ws.row_dimensions[lin].height = 18
        lin += 1

    # ── Linha de total geral ──────────────────────────────────────────────
    ws.merge_cells(f"A{lin}:G{lin}")
    ct = ws.cell(row=lin, column=1, value="TOTAL GERAL")
    ct.font      = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    ct.fill      = PatternFill("solid", fgColor=COR_HEADER)
    ct.alignment = Alignment(horizontal="right", vertical="center")

    total_cell = ws.cell(row=lin, column=8,
        value=f"=SUM(H5:H{lin - 1})")
    total_cell.fill         = PatternFill("solid", fgColor="D6EAF8")
    total_cell.font         = Font(name="Arial", bold=True, size=11, color="1E3A5F")
    total_cell.alignment    = Alignment(horizontal="right", vertical="center")
    total_cell.number_format = 'R$ #,##0.00'
    total_cell.border       = BORDA_FINA
    ws.row_dimensions[lin].height = 24

    # ── Rodapé ────────────────────────────────────────────────────────────
    lin += 2
    ws.merge_cells(f"A{lin}:I{lin}")
    rod = ws.cell(row=lin, column=1,
        value="* Proposta gerada automaticamente pelo sistema Projetista 360 — IceNexus IAR. "
              "Os preços desta planilha são de responsabilidade do fornecedor.")
    rod.font      = Font(name="Arial", size=7, color="999999", italic=True)
    rod.fill      = PatternFill("solid", fgColor=COR_RODAPE)
    rod.alignment = Alignment(horizontal="left", vertical="center")

    # ── Congelar cabeçalho ────────────────────────────────────────────────
    ws.freeze_panes = "A5"

    # ── Salvar em bytes ───────────────────────────────────────────────────
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()

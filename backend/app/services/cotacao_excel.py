"""
Gera planilha Excel de cotação vinculada a um fornecedor.

Estrutura de colunas:
  A  Código
  B  Tipo
  C  Descrição
  D  Detalhe / Esp. Técnica
  E  Qtde (m)   ← metros de tubo (informativo, apenas tubos de cobre)
  F  Qtde       ← quantidade comercial (kg para tubos, unidade normal para outros)
  G  Un
  H  Preço Unit. (R$)   ← fornecedor preenche
  I  Total (R$)          ← fórmula F×H
  J  Marca/Modelo        ← fornecedor preenche
  K  Prazo (dias)        ← fornecedor preenche
  L  Obs. Fornecedor     ← fornecedor preenche

Células travadas: tudo, exceto colunas que o fornecedor preenche: H, J, K, L
"""
from __future__ import annotations
from io import BytesIO
from datetime import datetime, timedelta

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.utils import get_column_letter

COR_HEADER    = "1E3A5F"
COR_EQUIP     = "EBF5FB"
COR_MATERIAL  = "FDFEFE"
COR_PREENCHER = "FFFACD"
COR_TOTAL     = "EAF2FF"
COR_RODAPE    = "F4F6F6"
COR_CODIGO    = "FFE8CC"
COR_METROS    = "F0FFF0"   # verde muito claro — coluna informativa de metros

BORDA_FINA = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

DESBLOQUEADA = Protection(locked=False)

# Colunas que o fornecedor preenche (índices 1-based)
# H=8 Preço Unit., J=10 Marca/Modelo, K=11 Prazo, L=12 Obs.
COLUNAS_EDITAVEIS = {8, 10, 11, 12}

COLUNAS = [
    ("Código",                  8),   # A=1
    ("Tipo",                   13),   # B=2
    ("Descrição",              40),   # C=3
    ("Detalhe / Esp. Técnica", 28),   # D=4
    ("Qtde (m)",                9),   # E=5 — metros de tubo (informativo)
    ("Qtde",                    8),   # F=6 — qtde comercial
    ("Un",                      6),   # G=7
    ("Preço Unit. (R$)",       15),   # H=8 ← fornecedor preenche
    ("Total (R$)",             13),   # I=9
    ("Marca/Modelo ofertado",  22),   # J=10 ← fornecedor preenche
    ("Prazo (dias)",           10),   # K=11 ← fornecedor preenche
    ("Obs. Fornecedor",        20),   # L=12 ← fornecedor preenche
]
ULT_COL = get_column_letter(len(COLUNAS))   # L


def _celula(ws, linha, col, valor, negrito=False, cor_fundo=None,
            cor_texto="000000", alinhamento="left", tamanho=10, italico=False):
    c = ws.cell(row=linha, column=col, value=valor)
    c.font = Font(name="Arial", bold=negrito, italic=italico,
                  color=cor_texto, size=tamanho)
    if cor_fundo:
        c.fill = PatternFill("solid", fgColor=cor_fundo)
    c.alignment = Alignment(horizontal=alinhamento, vertical="center")
    c.border = BORDA_FINA
    return c


def gerar_planilha_cotacao_v2(
    codigo: str,
    itens: list[dict],
    fornecedor_nome: str,
    fornecedor_contato: str = "",
    nome_projeto: str = "Projeto",
    nome_cliente: str = "",
    validade_dias: int = 30,
) -> bytes:
    """
    Cada item: {ref_id, tipo, descricao, detalhe, qtde, unidade, qtde_metros?}
    qtde_metros: presente apenas em tubos de cobre. qtde nesses casos está em kg.
    Retorna bytes do .xlsx.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Cotação"
    ws.sheet_view.showGridLines = False

    wb.properties.subject = codigo
    wb.properties.keywords = f"cotacao:{codigo}"

    # ── Cabeçalho ─────────────────────────────────────────────────────────
    ws.merge_cells(f"A1:{ULT_COL}1")
    c = ws.cell(row=1, column=1, value="PLANILHA DE COTAÇÃO — MATERIAIS E EQUIPAMENTOS")
    c.font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=COR_HEADER)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:C2")
    cc = ws.cell(row=2, column=1, value=f"CÓDIGO DA COTAÇÃO: {codigo}")
    cc.font = Font(name="Arial", bold=True, size=11, color="8B4513")
    cc.fill = PatternFill("solid", fgColor=COR_CODIGO)
    cc.alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells(f"D2:{ULT_COL}2")
    emissao = datetime.now()
    validade = emissao + timedelta(days=validade_dias)
    cd = ws.cell(
        row=2, column=4,
        value=(f"Fornecedor: {fornecedor_nome}"
               + (f" ({fornecedor_contato})" if fornecedor_contato else "")
               + f" | Emissão: {emissao.strftime('%d/%m/%Y')}"
               + f" | Validade solicitada: {validade.strftime('%d/%m/%Y')}"),
    )
    cd.font = Font(name="Arial", size=9, color="555555")
    cd.fill = PatternFill("solid", fgColor=COR_CODIGO)
    cd.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 20

    ws.merge_cells(f"A3:{ULT_COL}3")
    info = f"Projeto: {nome_projeto}"
    if nome_cliente:
        info += f" | Cliente: {nome_cliente}"
    c3 = ws.cell(row=3, column=1, value=info)
    c3.font = Font(name="Arial", size=9, color="555555", italic=True)
    c3.fill = PatternFill("solid", fgColor="EBF2FF")
    c3.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 16

    ws.merge_cells(f"A4:{ULT_COL}4")
    inst = ws.cell(
        row=4, column=1,
        value=("INSTRUÇÕES: Preencha apenas as colunas em amarelo — Preço Unitário, "
               "Marca/Modelo ofertado, Prazo e Observações. "
               "A coluna 'Qtde (m)' é informativa (metragem de tubo). "
               "O preço deve ser cotado por kg. "
               "NÃO altere o código da cotação."),
    )
    inst.font = Font(name="Arial", size=8, color="8B4513", bold=True)
    inst.fill = PatternFill("solid", fgColor=COR_PREENCHER)
    inst.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[4].height = 16

    # ── Cabeçalho de colunas ──────────────────────────────────────────────
    LINHA_HDR = 5
    for col_idx, (titulo, largura) in enumerate(COLUNAS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = largura
        editavel = col_idx in COLUNAS_EDITAVEIS
        metros_col = (col_idx == 5)  # coluna E — Qtde (m)
        c = ws.cell(row=LINHA_HDR, column=col_idx, value=titulo)
        if editavel:
            cor_hdr, txt_hdr = COR_PREENCHER, "000000"
        elif metros_col:
            cor_hdr, txt_hdr = "2E8B57", "FFFFFF"  # verde escuro
        else:
            cor_hdr, txt_hdr = COR_HEADER, "FFFFFF"
        c.font = Font(name="Arial", bold=True, size=9, color=txt_hdr)
        c.fill = PatternFill("solid", fgColor=cor_hdr)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDA_FINA
    ws.row_dimensions[LINHA_HDR].height = 26

    # ── Itens ─────────────────────────────────────────────────────────────
    lin = LINHA_HDR + 1
    primeira_linha_item = lin
    for item in itens:
        cor = COR_EQUIP if item.get("tipo") == "Equipamento" else COR_MATERIAL
        qtde_metros = item.get("qtde_metros")

        _celula(ws, lin, 1, str(item.get("ref_id") or "-"), cor_fundo=cor,
                alinhamento="center", tamanho=8, cor_texto="888888")
        _celula(ws, lin, 2, item.get("tipo", ""), cor_fundo=cor, tamanho=9)
        _celula(ws, lin, 3, item.get("descricao", ""), cor_fundo=cor, negrito=True, tamanho=9)
        _celula(ws, lin, 4, item.get("detalhe", ""), cor_fundo=cor, tamanho=8, cor_texto="555555")

        # E — Qtde (m): preenchido apenas para tubos de cobre
        if qtde_metros:
            ce = ws.cell(row=lin, column=5, value=round(qtde_metros, 2))
            ce.fill = PatternFill("solid", fgColor=COR_METROS)
            ce.border = BORDA_FINA
            ce.alignment = Alignment(horizontal="center", vertical="center")
            ce.font = Font(name="Arial", size=9, color="2E8B57", bold=True)
            ce.number_format = '#,##0.00'
        else:
            _celula(ws, lin, 5, "", cor_fundo=cor)

        # F — Qtde comercial (kg para tubos, normal para outros)
        _celula(ws, lin, 6, round(item.get("qtde", 1), 3), cor_fundo=cor,
                alinhamento="center", tamanho=9)

        # G — Unidade
        _celula(ws, lin, 7, item.get("unidade", "un"), cor_fundo=cor,
                alinhamento="center", tamanho=9)

        # H — preço unitário (editável)
        ch = ws.cell(row=lin, column=8)
        ch.fill = PatternFill("solid", fgColor=COR_PREENCHER)
        ch.border = BORDA_FINA
        ch.alignment = Alignment(horizontal="right", vertical="center")
        ch.number_format = "#,##0.00"
        ch.protection = DESBLOQUEADA

        # I — total: Qtde (F) × Preço (H)
        ci = ws.cell(row=lin, column=9, value=f'=IF(H{lin}="","",F{lin}*H{lin})')
        ci.fill = PatternFill("solid", fgColor=COR_TOTAL)
        ci.border = BORDA_FINA
        ci.alignment = Alignment(horizontal="right", vertical="center")
        ci.number_format = "#,##0.00"
        ci.font = Font(name="Arial", bold=True, size=9)

        # J, K, L — marca/modelo, prazo, obs (editáveis)
        for col in (10, 11, 12):
            cee = ws.cell(row=lin, column=col)
            cee.fill = PatternFill("solid", fgColor=COR_PREENCHER)
            cee.border = BORDA_FINA
            cee.alignment = Alignment(
                horizontal="center" if col == 11 else "left",
                vertical="center"
            )
            cee.protection = DESBLOQUEADA

        ws.row_dimensions[lin].height = 18
        lin += 1

    # ── Total geral ───────────────────────────────────────────────────────
    ws.merge_cells(f"A{lin}:H{lin}")
    ct = ws.cell(row=lin, column=1, value="TOTAL GERAL")
    ct.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    ct.fill = PatternFill("solid", fgColor=COR_HEADER)
    ct.alignment = Alignment(horizontal="right", vertical="center")

    tc = ws.cell(row=lin, column=9,
                 value=f"=SUM(I{primeira_linha_item}:I{lin - 1})")
    tc.fill = PatternFill("solid", fgColor="D6EAF8")
    tc.font = Font(name="Arial", bold=True, size=11, color="1E3A5F")
    tc.alignment = Alignment(horizontal="right", vertical="center")
    tc.number_format = 'R$ #,##0.00'
    tc.border = BORDA_FINA
    ws.row_dimensions[lin].height = 24

    # ── Rodapé ────────────────────────────────────────────────────────────
    lin += 2
    ws.merge_cells(f"A{lin}:{ULT_COL}{lin}")
    rod = ws.cell(
        row=lin, column=1,
        value=(f"* Cotação {codigo} gerada automaticamente. "
               "Os preços informados são de responsabilidade do fornecedor. "
               "Tubos de cobre: coluna 'Qtde (m)' é informativa — cotar e pagar por kg. "
               "Devolva este arquivo preenchido sem alterar sua estrutura."),
    )
    rod.font = Font(name="Arial", size=7, color="999999", italic=True)
    rod.fill = PatternFill("solid", fgColor=COR_RODAPE)
    rod.alignment = Alignment(horizontal="left", vertical="center")

    # ── Proteção da planilha ──────────────────────────────────────────────
    ws.protection.sheet = True
    ws.protection.formatColumns = False
    ws.protection.formatRows = False

    ws.freeze_panes = f"A{primeira_linha_item}"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()

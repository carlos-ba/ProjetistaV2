"""
Importa a planilha de cotação devolvida pelo fornecedor (Fase 2).

Estratégia defensiva — o fornecedor pode ter alterado a planilha:
  1. Identifica a cotação pelo código (propriedade do documento OU célula A2)
  2. Casa as linhas pelo par (nº da linha, descrição) com os itens do banco
  3. Tolera: linhas removidas, preço vazio, texto no lugar de número
  4. Nada é gravado aqui — apenas retorna o que foi lido para conferência
"""
from __future__ import annotations
import re
from io import BytesIO
from typing import Any

from openpyxl import load_workbook

# Colunas da planilha v2 (1-based)
COL_REF, COL_TIPO, COL_DESC, COL_DET, COL_QTDE, COL_UN = 1, 2, 3, 4, 5, 6
COL_PRECO, COL_TOTAL, COL_MARCA, COL_PRAZO, COL_OBS = 7, 8, 9, 10, 11

LINHA_PRIMEIRO_ITEM = 6   # linha 5 é o cabeçalho das colunas

_RE_CODIGO = re.compile(r"COT-\d{4}-\d{4}-F\d{3}")


def extrair_codigo(conteudo: bytes) -> str | None:
    """Lê o código da cotação: propriedades do documento → célula A2 → varredura."""
    wb = load_workbook(BytesIO(conteudo), read_only=True, data_only=True)
    try:
        # 1. Propriedades do documento
        for prop in (wb.properties.subject, wb.properties.keywords):
            if prop:
                m = _RE_CODIGO.search(str(prop))
                if m:
                    return m.group(0)

        # 2. Célula A2 (cabeçalho visível) e varredura das primeiras linhas
        ws = wb.active
        for row in ws.iter_rows(min_row=1, max_row=6, max_col=4):
            for cell in row:
                if cell.value:
                    m = _RE_CODIGO.search(str(cell.value))
                    if m:
                        return m.group(0)
        return None
    finally:
        wb.close()


def _parse_numero(valor: Any) -> float | None:
    """Converte o que o fornecedor digitou em número — tolera 'R$ 1.234,56', '1234.56', etc."""
    if valor is None:
        return None
    if isinstance(valor, (int, float)):
        return float(valor)
    s = str(valor).strip()
    if not s:
        return None
    s = re.sub(r"[Rr]\$\s*", "", s)
    # Formato brasileiro: 1.234,56 → 1234.56
    if "," in s:
        s = s.replace(".", "").replace(",", ".")
    s = re.sub(r"[^\d.\-]", "", s)
    try:
        return float(s)
    except ValueError:
        return None


def _parse_inteiro(valor: Any) -> int | None:
    n = _parse_numero(valor)
    return int(n) if n is not None else None


def ler_itens_planilha(conteudo: bytes) -> list[dict]:
    """
    Lê as linhas de itens da planilha devolvida.
    Retorna lista de dicts com o que o fornecedor preencheu (sem gravar nada).
    """
    wb = load_workbook(BytesIO(conteudo), read_only=True, data_only=True)
    try:
        ws = wb.active
        itens: list[dict] = []
        for row in ws.iter_rows(min_row=LINHA_PRIMEIRO_ITEM, max_col=COL_OBS):
            desc = row[COL_DESC - 1].value
            # Fim da tabela: linha TOTAL GERAL ou vazia
            primeira = str(row[0].value or "")
            if "TOTAL" in primeira.upper():
                break
            if not desc or not str(desc).strip():
                continue

            itens.append({
                "linha": row[0].row,
                "ref_id": _parse_inteiro(row[COL_REF - 1].value),
                "tipo": str(row[COL_TIPO - 1].value or "").strip(),
                "descricao": str(desc).strip(),
                "qtde": _parse_numero(row[COL_QTDE - 1].value),
                "unidade": str(row[COL_UN - 1].value or "").strip(),
                "preco_unitario": _parse_numero(row[COL_PRECO - 1].value),
                "marca_modelo": (str(row[COL_MARCA - 1].value).strip()
                                 if row[COL_MARCA - 1].value else None),
                "prazo_dias": _parse_inteiro(row[COL_PRAZO - 1].value),
                "obs": (str(row[COL_OBS - 1].value).strip()
                        if row[COL_OBS - 1].value else None),
                "preco_bruto": (str(row[COL_PRECO - 1].value)
                                if row[COL_PRECO - 1].value is not None else None),
            })
        return itens
    finally:
        wb.close()


def casar_itens(itens_banco: list, itens_planilha: list[dict]) -> list[dict]:
    """
    Casa os itens lidos da planilha com os itens da cotação no banco.

    Estratégia: descrição exata → descrição normalizada → sem correspondência.
    Retorna um relatório por item do banco + linhas extras da planilha.
    """
    def norm(s: str) -> str:
        return re.sub(r"\s+", " ", str(s or "").strip().lower())

    usados: set[int] = set()
    relatorio: list[dict] = []

    for item_db in itens_banco:
        encontrado = None
        # 1. Descrição exata
        for idx, ip in enumerate(itens_planilha):
            if idx in usados:
                continue
            if norm(ip["descricao"]) == norm(item_db.descricao):
                encontrado = (idx, ip)
                break

        if encontrado:
            idx, ip = encontrado
            usados.add(idx)
            tem_preco = ip["preco_unitario"] is not None and ip["preco_unitario"] > 0
            status = "ok" if tem_preco else "sem_preco"
            # Preço com texto que não converteu → alerta
            if not tem_preco and ip["preco_bruto"]:
                status = "preco_invalido"
            relatorio.append({
                "item_id": item_db.id,
                "descricao": item_db.descricao,
                "qtde": float(item_db.qtde),
                "unidade": item_db.unidade,
                "status": status,
                "preco_unitario": ip["preco_unitario"],
                "preco_bruto": ip["preco_bruto"],
                "marca_modelo": ip["marca_modelo"],
                "prazo_dias": ip["prazo_dias"],
                "obs": ip["obs"],
            })
        else:
            relatorio.append({
                "item_id": item_db.id,
                "descricao": item_db.descricao,
                "qtde": float(item_db.qtde),
                "unidade": item_db.unidade,
                "status": "nao_encontrado",
                "preco_unitario": None,
                "preco_bruto": None,
                "marca_modelo": None,
                "prazo_dias": None,
                "obs": None,
            })

    # Linhas da planilha que não casaram com nenhum item do banco
    extras = [
        {
            "item_id": None,
            "descricao": ip["descricao"],
            "qtde": ip["qtde"] or 1,
            "unidade": ip["unidade"] or "un",
            "status": "linha_extra",
            "preco_unitario": ip["preco_unitario"],
            "preco_bruto": ip["preco_bruto"],
            "marca_modelo": ip["marca_modelo"],
            "prazo_dias": ip["prazo_dias"],
            "obs": ip["obs"],
        }
        for idx, ip in enumerate(itens_planilha)
        if idx not in usados
    ]

    return relatorio + extras
